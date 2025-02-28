# author:Soufaker
# time:2023/09/05
import random
import traceback
import requests
import time
import optparse
import json
import base64
import urllib.request
import re
import tldextract
from dingtalkchatbot.chatbot import DingtalkChatbot
from datetime import datetime
import math
from openpyxl import Workbook
from configparser import ConfigParser

requests.packages.urllib3.disable_warnings()
import urllib
import os


def beianchaxun(url):
    urls = url
    URL = ("https://beian.tianyancha.com/search/{}".format(urls))
    headers = {
        'Connection': 'close',
        'Cache-Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.8',
    }
    try:
        req = requests.get(url=URL, headers=headers)
        html = req.content
        titlere = r'<span class="ranking-ym" rel="nofollow">(.+?)</span>'
        title = re.findall(titlere, html.decode('utf-8'))
        if len(title) == 0:
            print(urls, "该公司天眼查找不到备案域名", '\n')
        else:
            print('\n', urls, "备案域名: ")
            for i in title:
                if i not in black_domian and i != '':
                    all_domain_list.append(i)
                    print(i)
    except:
        print('有异常，无法爬取')


# 选取想要的域名元素
def Get_MiddleStr(content, startStr, endStr, num):  # 获取中间字符串的⼀个通⽤函数
    try:
        startIndex = content.index(startStr)
        if startIndex >= 0:
            if num != 1:
                endIndex = startIndex
                startIndex = endIndex + num

            else:
                startIndex += len(startStr)
                content2 = content[startIndex:]
                endIndex = content2.index(endStr) + startIndex

        return content[startIndex:endIndex]

    except:
        return 'ok'


# 获取公司查询ID
def Get_Company_Id(company_name, company_name_list):
    company_id_name_list = []
    if company_name != None:
        com_id = []
        id = get_company_jt_info(company_name)[0]
        name = get_company_jt_info(company_name)[1]
        com_id.append(id)
        com_id.append(name.replace('<em>', '').replace('</em>', ''))
        company_id_name_list.append(com_id)

    else:
        with open(company_name_list, 'r', encoding='utf-8') as f:
            t = f.readlines()
            for i in t:
                com_id = []
                id = get_company_jt_info(i.strip('\n'))[0]
                name = get_company_jt_info(i.strip('\n'))[1]
                com_id.append(id)
                com_id.append(name.replace('<em>', '').replace('</em>', ''))
                company_id_name_list.append(com_id)

    return company_id_name_list


# 获取公司占百分百股权且未注销的子公司域名信息
def Get_ALL_Sub_Cmpany_Domain(company_id_name_list):
    company_info_list = []
    company_url_list = []
    com_id = []
    for l in company_id_name_list:
        company_info = get_company_jt_info(l[1])
        info = company_info[1:]
        company_info_list.append(info)
        company_url_list.append(info[2])
        com_id.append(l[0])

    while com_id:
        co_id = com_id.pop(0)
        company_id_name_list = get_all_page_id(co_id)
        if len(company_id_name_list) != 0:
            for co in company_id_name_list:
                com_id.append(co[0])
            for ci in company_id_name_list:
                try:
                    company_info = get_company_jt_info(ci[1])
                    info = company_info[1:]
                    company_info_list.append(info)
                    company_url_list.append(info[2])
                except:
                    continue

    return company_info_list, company_url_list


def Get_Sub_Company_Domain(company_id):
    company_info_list = get_company_jt_info(company_id)
    info = company_info_list[1:]
    company_url_list = company_info_list[2]

    return info, company_url_list


def get_all_page_id(id):
    id_list = []
    company_url = "https://capi.tianyancha.com/cloud-company-background/company/investListV2?_=1663738376979"
    data = '{"gid":"' + str(id) + '","pageSize":100,"pageNum":1,"province":"","percentLevel":"-100","category":"-100"}'
    header = {
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    header['X-AUTH-TOKEN'] = x_auth_token
    try:
        response = requests.post(url=company_url, headers=header, data=data).text
        j = json.loads(response)
        print(j)
        for i in range(len(j['data']['result'])):
            l1 = []
            occ_r = j['data']['result'][i]['percent'][:-1]
            if occ_r != '':
                occ_rev = float(occ_r)
                if occ_rev >= float(company_occ):
                    l1.append(j['data']['result'][i]['id'])
                    l1.append(j['data']['result'][i]['name'])
                    # l1.append(j['data']['result'][i]['regStatus'])
                    id_list.append(l1)

    except:
        print('IP地址可能被ban。。请切换IP或者稍后再试!')
        return id_list

    return id_list


def Write_To_Excel(company_info_list, all_info_list, mgwj_list, ld_list, httpx_info,fs_list):
    t = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
    wb = Workbook()
    ws1 = wb.create_sheet('天眼查基本信息', 0)
    ws2 = wb.create_sheet('收集域名信息', 0)
    ws3 = wb.create_sheet('fscan扫描信息',0)
    ws4 = wb.create_sheet('敏感信息', 0)
    ws5 = wb.create_sheet('漏洞信息', 0)
    ws6 = wb.create_sheet('httpx信息', 0)


    ws1['A1'] = '公司名'
    ws1['B1'] = '网址'
    ws1['C1'] = '邮箱'
    ws1['D1'] = '联系电话'
    for l in company_info_list:
        ws1.append(list(l))

    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 36
    ws1.column_dimensions['C'].width = 36
    ws1.column_dimensions['D'].width = 36

    ws2['A1'] = '网址'
    ws2['B1'] = 'ip地址'
    ws2['C1'] = '端口'
    ws2['D1'] = '网页标题'
    ws2['E1'] = '协议'
    ws2['F1'] = '状态码'
    ws2['G1'] = '框架'
    ws2['H1'] = '备案号'
    for l in all_info_list:
        ws2.append(l)

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 7
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 7
    ws2.column_dimensions['F'].width = 7
    ws2.column_dimensions['G'].width = 30
    ws2.column_dimensions['H'].width = 25

    ws3['A1'] = '漏洞地址'
    ws3.column_dimensions['A'].width = 70
    for l in fs_list:
        list1 = []
        list1.append(l)
        ws3.append(list1)

    ws4['A1'] = '地址'
    ws4['B1'] = '状态码'
    ws4['C1'] = '大小'

    ws4.column_dimensions['A'].width = 70
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 15

    for l in mgwj_list:
        ws4.append(l)

    ws5['A1'] = '漏洞名字'
    ws5['B1'] = '漏洞等級'
    ws5['C1'] = '漏洞地址'

    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 10
    ws5.column_dimensions['C'].width = 50

    for l in ld_list:
        ws5.append(l)

    ws6['A1'] = '网址'
    ws6['B1'] = '状态码'
    ws6['C1'] = '标题'

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 50

    for l in httpx_info:
        ws6.append(l)


    wb.save("./result/baolumian/暴露面收集" + t + ".xlsx")


def Get_icp_Num(company_name):
    company_url = "https://beian.tianyancha.com/search/" + company_name
    print(company_url)
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    try:
        response = requests.get(url=company_url, headers=header).text
        resp = response.replace(' ', '').replace('\n', '').replace('\t', '').replace("\"", "")

        flag = True
        page_num = 1

        while flag:
            for i in range(2, 100):
                str_flag = '/p' + str(i)
                if str_flag in resp:
                    flag = True
                    page_num += 1
                else:
                    flag = False
    except:
        page_num = 0

    return page_num


def yt_info(url):
    temp_url_list = []
    temp_url_list.append(url)
    while temp_url_list:
        temp_url_list.pop()
        r = requests.get(url, timeout=3)
        res = json.loads(r.text)
        if str(res['code']) == '429':
            temp_url_list.append(url)
            time.sleep(1)
            continue
        loadurl = res['data']['arr']

        if loadurl != None:
            company = loadurl[0]['company']
            all_company_name_list.append(company)
            print(company)
            for i in range(0, len(loadurl)):
                domain = loadurl[i]['domain']
                if '.' in domain:
                    val = tldextract.extract(domain)
                    if val.registered_domain not in all_domain_list:
                        continue
                info = []
                arr_url = loadurl[i]['url']
                # if '中国' in loadurl[i]['isp']:
                #     arr_ip = loadurl[i]['ip']
                # else:
                #     arr_ip = '存在CDN:' + loadurl[i]['ip']
                arr_ip = isCDN(loadurl[i]['domain'], loadurl[i]['ip'])
                arr_port = loadurl[i]['port']
                arr_web_title = loadurl[i]['web_title']
                arr_protocol = loadurl[i]['protocol'] + ',' + loadurl[i]['base_protocol']
                arr_status_code = loadurl[i]['status_code']
                arr_component = loadurl[i]['component']
                arr_beianhhao = loadurl[i]['number']
                if arr_beianhhao == '':
                    arr_beianhhao = '-'
                arr_all_component = ''
                if arr_component != None:
                    for i in arr_component:
                        s = i['name'] + i['version']
                        arr_all_component = arr_all_component + '|' + s
                info.append(arr_url)
                info.append(arr_ip)
                info.append(arr_port)
                info.append(arr_web_title)
                info.append(arr_protocol)
                info.append(arr_status_code)
                info.append(arr_all_component)
                info.append(arr_beianhhao)
                # print(info)
                flag = True
                for b in black_domian:
                    if b in arr_url:
                        flag = False
                if flag:
                    all_info_list.append(info)
            return 1
        else:
            return 2


def yt_get_info(name_list):
    new_list = fy_list(name_list, hunter_count)
    for domain_list in new_list:
        try:
            domain_all = ''
            for domain in domain_list:
                if domain != '':
                    if isIP(domain):
                        domain_all = domain_all + "ip=" + domain + '||'
                    else:
                        domain_all = domain_all + "domain=" + domain + '||'
            print(domain_all)
            search_key = '(' + domain_all[0:-2] + ')' + str(fofa_keyword)
            keyword = base64.urlsafe_b64encode(search_key.encode("utf-8"))  # 把输入的关键字转换为base64编码
            page = 1
            api_num = 0
            while True:
                # 测试第一个API积分是否够用
                url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size=1&is_web=1".format(
                    hunter_config_list[api_num], keyword.decode(), page)
                r = requests.get(url)
                res = json.loads(r.text)

                if str(res['code']) == '429':
                    continue
                if str(res['code']) == '401':
                    if int(api_num) < int(len(hunter_config_list)):
                        api_num += 1
                    continue
                if str(res['code']) == '40204' or str(res['code']) == '40201':
                    if int(api_num) < int(len(hunter_config_list)):
                        api_num += 1
                        print('上一个积分已经用完,切换第' + str(int(api_num) + 1) + '个API')
                        continue

                url = "https://hunter.qianxin.com/openApi/search?api-key={}&search={}&page={}&page_size={}&is_web=1".format(
                    hunter_config_list[api_num], keyword.decode(), page, yt_size)
                print(url)
                pd_num = yt_info(url)
                if pd_num == 2:
                    print('未查到数据')
                break
        except Exception as e:
            traceback.print_exc()
            print('积分清0退出循环', e)
            break


def get_title(url):
    try:
        page = urllib.request.urlopen(url=url, timeout=0.5)
        html = page.read().decode('utf-8')
        title = re.findall('<title>(.+)</title>', html)
        return title[0]
    except:
        return ''


def isCDN(domain, ip):  # 判断目标是否存在CDN
    parm = 'nslookup ' + domain
    try:
        result = os.popen(parm).read()
        l = result.split('Name:')
        if result.count("Name") > 1 or domain not in l[1]:
            return "存在CDN" + str(ip)
        else:
            if ip not in ip_list:
                ip_list.append(ip)
            return ip
    except:
        return '-'

def fy_list(list1, count):
    new_list = []
    num = len(list1)
    if int(num) > int(count):
        n = num // int(count)
    else:
        n = 1

    for i in range(0, n):
        one_list = list1[math.floor(i / n * num):math.floor((i + 1) / n * num)]
        new_list.append(one_list)
    return new_list


def get_fofa_url(domain_l):
    new_list = fy_list(domain_l, fofa_count)
    while True:
        for domain_list in new_list:
            try:
                domain_all = ''
                for domain in domain_list:
                    if domain != '':
                        if isIP(domain):
                            domain_all = domain_all + "ip=" + domain + '||'
                        else:
                            domain_all = domain_all + "domain=" + domain + '||'
                print(domain_all)
                search_key = '(' + domain_all[0:-2] + ')' + str(fofa_keyword)
                search_data_b64 = base64.b64encode(search_key.encode("utf-8")).decode("utf-8")
                search = 'https://fofa.info/api/v1/search/all?email=' + fofa_email + '&size=' + fofa_size + '&key=' + fofa_key + '&qbase64=' + search_data_b64 + "&fields=host,ip,port,title,protocol,header,server,product,icp,domain"
                print(search)
                try:
                    r = requests.get(search, verify=False)
                    res = json.loads(r.text)
                    size = len(res['results'])
                    for i in range(0, int(size)):
                        info = []
                        result = res['results'][i]
                        num = len(result)
                        temp = ''
                        if result[4] == 'unknown':
                            result[4] = 'http'
                        if result[9] not in domain_l:
                            continue
                        for j in range(0, int(num) - 1):
                            if j == 1:
                                ip = isCDN(result[9], result[1])
                                result[1] = ip
                            if j == 0:
                                print(result[0])
                                if 'http' not in result[0][0:5]:
                                    if 'http' not in result[4]:
                                        result[0] = 'https://' + result[0]
                                    else:
                                        result[0] = result[4] + '://' + result[0]
                            if j == 6:
                                temp = result[j]
                                continue
                            if j == 7:
                                result[j] = '|' + result[j] + '|' + temp
                            if j == 8:
                                if result[j] == '':
                                    result[j] = '-'
                            if j == 5:
                                if result[j] == '':
                                    result[j] == '-'
                            if j == 3:
                                result[j] = get_title(result[0])
                                if result[j] == '':
                                    result[j] = '-'

                            if j == 5:
                                result[j] = result[j][9:12]
                            info.append(result[j])
                        print(info)
                        flag = True
                        for b in black_domian:
                            if b in result[0]:
                                flag = False
                        if flag:
                            all_info_list.append(info)
                except:
                    continue
            except:
                print('fofa连接超时,正在重试！')
                time.sleep(1)
                continue
        print('查询结束了')
        break


def split_list_average_n(origin_list, n):
    for i in range(0, len(origin_list), n):
        yield origin_list[i:i + n]


def get_all_url_fo_yt(company_info_list, company_domains_file):
    company_domains_list = []
    if len(company_domains_file) > 2:
        with open(company_domains_file, 'r', encoding='utf-8') as f:
            l = f.readlines()
            for i in l:
                company_domains_list.append(i.strip('\n'))

    # 查询所有域名
    for com in company_info_list:
        # 备案查询域名
        beianchaxun(com[0])

        # 公司备案信息查询的域名
        company_info_list = com[1].split(',')
        for com in company_info_list:
            if '.' in com:
                val = tldextract.extract(com)
                if val.registered_domain != '':
                    all_domain_list.append(val.registered_domain)

    if len(company_domains_list) != 0:
        for com in company_domains_list:
            all_domain_list.append(com)

    print('当前搜集了如下域名')
    all_qc_domain_list = list(set(all_domain_list))
    filename = './result/alldomain/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_all_domain_list.txt'
    with open(filename, 'w', encoding='utf-8') as f:
        for a in all_qc_domain_list:
            f.writelines(a + '\n')
    print(all_qc_domain_list)
    if x_domain != '1':
        if notauto != True:
            # 调用鹰图,并添加到所有搜集的列表
            if is_hunter == '0':
                print('开始调用鹰图')
                yt_get_info(all_qc_domain_list)
            if is_fofa == '0':
                print('开始使用fofa查询')
                get_fofa_url(all_qc_domain_list)


def get_company_jt_info(name):
    all_info = []
    # proxies = {'http': 'http://localhost:8080', 'https': 'http://localhost:8080'}

    company_url = "https://capi.tianyancha.com/cloud-tempest/web/searchCompanyV3?_=1672969688987"
    data = '{"word":"' + str(name) + '","sortType":"1","pageSize":1,"referer":"search","pageNum":1}'
    print(data)
    header = {
        "host": "capi.tianyancha.com",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0",
        "version": "TYC-Web",
    }
    header['X-AUTH-TOKEN'] = x_auth_token
    try:
        response = requests.post(url=company_url, headers=header, data=data.encode('utf-8'), verify=False).text
        j = json.loads(response)
        id = j['data']['companyList'][0]['id']
        name = j['data']['companyList'][0]['name']
        email = j['data']['companyList'][0]['emailList']
        phone = j['data']['companyList'][0]['phoneList']
        site = str(j['data']['companyList'][0]['websites']).split('\t')
        website = []
        for w in site:
            if ';' not in w:
                website.append(w)

        all_info.append(id)
        all_info.append(name.replace('<em>', '').replace('</em>', ''))
        if website != None:
            all_info.append(','.join(map(str, website)))
        else:
            all_info.append('无信息')

        if email != None:
            all_info.append(','.join(map(str, email)))
        else:
            all_info.append('无信息')

        if phone != None:
            all_info.append(','.join(map(str, phone)))
        else:
            all_info.append('无信息')
    except Exception as e:
        print('出异常了', e)

    return all_info

#判断是ip还是域名
def isIP(str):
    p = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if p.match(str):
        return True
    else:
        return False

def save_cache(target_list):
    yt_fofa_add_list = []
    yt_fofa_add_list2 = []
    sm_add_list = []
    # 读取历史扫描信息
    sm_cache_file_list = open('./caches/sm_cache.txt', 'r', encoding='utf-8').read().split('\n')
    #print('target', target_list)
    #print('sm', sm_cache_file_list)
    # 添加fafa_yt记录的历史域名信息
    if len(target_list) != 0:
        for tar in target_list:
            print(tar[0])
            if tar[0] not in sm_cache_file_list:
                if str(tar[5]) == '200' or str(tar[5]) == '301' or str(tar[5]) == '302' or str(tar[5]) == '201' or str(tar[5]) == '404' or str(tar[5]) == '401' or str(tar[5]) == '405':
                    info = []
                    info.append(str(tar[0]))
                    info.append(str(tar[5]))
                    info.append(str(tar[3]))
                    httpx_info.append(info)

                sm_add_list.append(tar[0])
                str_tar = ''
                for t in tar:
                    str_tar = str_tar + ' | ' + str(t)
                yt_fofa_add_list.append(str_tar)
                yt_fofa_add_list2.append(tar)

    for l in yt_fofa_add_list:
        caches_file = open('./caches/fo_yt_cache.txt', 'a', encoding='utf-8')
        caches_file.write(l + '\n')
        caches_file.close()

    return sm_add_list, yt_fofa_add_list2, sm_cache_file_list


def httpx_naabu_scan(filename, sm_cache_file_list):
    caches_file_list = open('./caches/sm_cache.txt', 'r', encoding='utf-8').read().split('\n')
    caches_file = open('./caches/sm_cache.txt', 'a', encoding='utf-8')
    try:
        file_list2 = open(filename, 'r', encoding='utf-8').read().split('\n')
        #print('-------------')
        #print(file_list2)
        filename_temp = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'temp_port.txt'
        temp_l = open(filename, 'r', encoding='utf-8').read().split('\n')
        #print(temp_l)
        new_filename_temp = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime()) + 'new_temp_port.txt'
        with open(new_filename_temp, 'w') as f:
            for i in temp_l:
                if ':' in i:
                    #print('123')
                    #print(i.split(':')[0])
                    f.writelines(i.split(':')[0] + '\n')
                else:
                    f.writelines(i+ '\n')

        filename_filter_name = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",
                                                                  time.localtime()) + 'all_url_list.txt'

        port_scan = './inifile/naabu/naabu  -l ' + new_filename_temp + ' -top-ports 1000 -o ' + filename_temp
        #print('2 ' + port_scan)
        os.system(port_scan)  # &> /dev/null
        httpx_filename = filename_temp[0:-4] + '_httpx.txt'
        http_list = open(filename_temp, 'r')
        #print('123'+httpx_filename)
        with open(httpx_filename, 'w+') as f:
            for h in http_list:
                #print(h)
                f.writelines('https://' + h)
                f.writelines('http://' + h)
        http_scan = './inifile/httpx/httpx  -l ' + httpx_filename + ' -mc 200,401,403,404,302,301  -title   -status-code  -fr -o  ' + filename_filter_name  # &> /dev/null'
        #print('1 ' + http_scan)
        os.system(http_scan)  # &> /dev/null
        # os.system('rm -rf ' + filename)
        # os.system('rm -rf ' + filename_temp)
        httpx_info_l = open(filename_filter_name, 'r', encoding='utf-8', errors='ignore').read().split('\n')
        httpx_info_list = list(set(httpx_info_l))
        for i in httpx_info_list:
            if i not in caches_file_list and i != '':
                info = []
                f = i.split(' ')
                info.append(str(f[0]))
                if '200' in f[1]:
                    info.append('200')
                else:
                    info.append(str(f[1].split('\x1b')[1].split('m')[1]))
                info.append(str(f[2].split('\x1b')[1][4:]))
                httpx_info.append(info)
                print(info)
        print(httpx_info)

        #
        new_filename_filter_name = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime()) + 'new_all_url_list.txt'
        print('new_filename_filter_name', new_filename_filter_name)
        caches_file_list_1 = open(new_filename_filter_name, 'w+', encoding='utf-8')
        # 写入awvs文件
        file_list = []
        for f in httpx_info:
            if f[0] not in file_list:
                file_list.append(f[0])
        print(file_list)
        #print(sm_cache_file_list)
        for l in file_list:
            if 'http' not in l:
                l1 = 'http://' + l
                l2 = 'https://' + l
                if l2 not in sm_cache_file_list:
                    caches_file.writelines(l2 + '\n')
                    caches_file_list_1.writelines(l2 + '\n')
                if l1 not in sm_cache_file_list:
                    caches_file.writelines(l1 + '\n')
                    caches_file_list_1.writelines(l1 + '\n')
            else:
                if l not in sm_cache_file_list:
                    print('0000000000000000000000')
                    caches_file.writelines(l + '\n')
                    caches_file_list_1.writelines(l + '\n')
        caches_file.close()
        # awvs
        scan_awvs(file_list)

        return new_filename_filter_name
    except Exception as e:
        traceback.print_exc()
        print(e)
        os.system('touch ' + filename_filter_name)
        print('12312312')
        if len(httpx_info) > 0:
            file_list2 = []
            for f in httpx_info:
                file_list2.append(f[0])
            print(file_list2)
            scan_awvs(file_list2)
            caches_file_list_1 = open(filename_filter_name, 'w', encoding='utf-8')
            for l in file_list2:
                if 'http' not in l:
                    l1 = 'http://' + l
                    l2 = 'https://' + l
                    if l2 not in sm_cache_file_list:
                        caches_file.write(l2 + '\n')
                        caches_file_list_1.write(l2 + '\n')
                    if l1 not in sm_cache_file_list:
                        caches_file.write(l1 + '\n')
                        caches_file_list_1.write(l1 + '\n')
                else:
                    if l not in sm_cache_file_list:
                        print('0000000000000000000000')
                        caches_file.write(l + '\n')
                        caches_file_list_1.write(l + '\n')

        return filename_filter_name


def scan_awvs(file_list):
    if avsm == True:
        print('开始调用awvs')
        os.system('nohup python3 awvs_monitor.py >awvsput.out 2>&1 &')
    awvs_file_name = './result/awvslist/all_av_list.txt'
    with open(awvs_file_name, 'a', encoding='utf-8') as f:
        for a in file_list:
            if 'http' in a:
                print(a)
                f.writelines(a + '\n')


def quchong_info_list(all_info_list):
    new_list = []
    mgwj_list = []
    ld_list = []
    fs_list = []
    for all in all_info_list:
        if all not in new_list:
            new_list.append(all)
    add_list, yt_fofa_info_list, sm_cache_file_list = save_cache(new_list)
    add_list = list(set(add_list))
    if len(add_list) != 0:
        filename = './result/allurl/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'all_url_list.txt'
        temp_list = []
        for a in add_list:
            if len(a) < 4:
                continue
            print('1' + a)
            try:
                a1 = a.split('://')[1]
                print(a1)
                if ':' in a1:
                    a1 = a1.split(':')[0]
                if '/' in a1:
                    temp_list.append(a1.split('/')[0])
                else:
                    temp_list.append(a1)
            except:
                continue
        temp_list2 = list(set(temp_list))
        with open(filename, 'w', encoding='utf-8') as f:
            for s in temp_list2:
                f.writelines(s+'\n')

        file_filter_name = httpx_naabu_scan(filename, sm_cache_file_list)
        print('ssss')
        print(file_filter_name)
        print('xxxxx')

        if ml == True:
            mgwj_list = ml_sm(file_filter_name)


        if fs == True:
            fs_list = fscan(file_filter_name,ip_list)

        if ld == True:
            ld_list = nuclei(file_filter_name)


    # 扫描自己收集的资产
    if notauto == True:
        filename = './result/notautolist/notautolist.txt'
        new_filename = './result/notautolist/'+str(time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()))+'new_notautolist.txt'
        add_list = open(filename, 'r', encoding='utf-8').read().split('\n')
        temp_list = []
        for a in add_list:
            if len(a) < 4:
                continue
            print('1' + a)
            try:
                a1 = a.split('://')[1]
                print(a1)
                if ':' in a1:
                    a1 = a1.split(':')[0]
                if '/' in a1:
                    temp_list.append(a1.split('/')[0])
                else:
                    temp_list.append(a1)

            except:
                if 'http:' not in a and 'htts:' not in a:
                    if ':' in a:
                        a = a.split(':')[0]
                    if '/' in a:
                        temp_list.append(a.split('/')[0])
                    else:
                        temp_list.append(a)
                continue
        temp_list2 = list(set(temp_list))

        with open(new_filename, 'w', encoding='utf-8') as f:
            for s in temp_list2:
                f.writelines(s+'\n')

        file_filter_name = httpx_naabu_scan(new_filename, sm_cache_file_list)
        print('ssss')
        print(file_filter_name)
        print('xxxxx')
        if ml == True:
            mgwj_list = ml_sm(file_filter_name)

        if ld == True:
            ld_list = nuclei(file_filter_name)

        if fs == True:
            fs_list = fscan(file_filter_name,ip_list)

    print('==============================')
    # print(mgwj_list)
    # print(ld_list)
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    print(yt_fofa_info_list)
    print('zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz')
    return yt_fofa_info_list, mgwj_list, ld_list,fs_list

def bypass403(mg_url):
    os.system('./inifile/bypass403/f403_linux_amd64 -u ' + mg_url + '> temp.txt')
    l = open('./inifile/bypass403/temp.txt', 'r', encoding='utf-8').read().split('\n')
    l2 = []
    for i in l:
        print(i)
        if '[*]' in i or '[+]' in i:
            l2.append(i)

    pass_url_list = []
    for i in range(0, len(l2)):
        print(l2[i])
        if '[+]' in l2[i]:
            info = []
            url = l2[i].split(' ')[3].split('\x1b')[0] + '==>' + \
                  l2[i - 1].replace(' ', '').split('[*]')[1].split('\x1b')[0]
            code = l2[i].split(' ')[1]
            words = l2[i].split(' ')[2]
            info.append(url)
            info.append(code)
            info.append(words)
            pass_url_list.append(info)

    os.system('rm -rf ./inifile/bypass403/temp.txt')
    return pass_url_list

def ml_sm(filename):
    url_list = open(filename, 'r', encoding='utf-8', errors='ignore').read().split('\n')
    # 返回的字节长度列表
    result = []

    for url in url_list:
        msg_info = []
        msg_info2 = []
        print(url)
        try:
            if 'http' in url:
                temp_file = 'temp_result.txt'
                print('./inifile/ffuf/ffuf -u ' + url + '/FUZZ -w ./inifile/dict/file_top_200.txt -ac -t 100 -o ' + temp_file)
                os.system('./inifile/ffuf/ffuf -u ' + url + '/FUZZ -w ./inifile/dict/file_top_200.txt -ac -t 100 -o ' + temp_file)

            else:
                continue


            with open(temp_file, 'r', encoding='utf-8') as f:
                data = json.load(f)['results']
                f.close()

            #删除临时文件
            os.system('rm -rf '+temp_file)
            #print(data)

            # 存放返回包长度
            for i in range(len(data)):
                msg_info.append(str(tldextract.extract(data[i]['url']).registered_domain) + str(data[i]['words']))
            for i in range(len(data)):
                msg_info2.append(str(data[i]['words']))
            #print(msg_info)

            for i in range(len(data)):
                info_list = []
                if msg_info.count(str(tldextract.extract(data[i]['url']).registered_domain) + str(data[i]['words'])) == 1 and data[i]['words'] > 100 and msg_info2.count(str(data[i]['words'])) < 4 :
                    info_list.append(data[i]['url'])
                    info_list.append(data[i]['status'])
                    info_list.append(data[i]['words'])
                    result.append(info_list)
                else:
                    print('1')

                #对扫描的403,401页面进行bypass扫描
                if data[i]['status'] == 403 or data[i]['status'] == 401:
                    l = bypass403(data[i]['url'])
                    if len(l) != 0:
                        for i in l:
                            result.append(i)
        except:
            continue

    return result


def dingtalk(message_list, mgml_list, ld_list, fs_list):
    # 钉钉WebHook地址
    DWebHook = 'https://oapi.dingtalk.com/robot/send?access_token=' + str(dingding_hook)
    Dsecret = dingding_key  # 可选：创建机器人勾选“加签”选项时使用
    # 初始化机器人小丁
    # xiaoding = DingtalkChatbot(webhook)  # 方式一：通常初始化方式
    xiaoding = DingtalkChatbot(DWebHook, secret=Dsecret)  # 方式二：勾选“加签”选项时使用（v1.5以上新功能）
    # xiaoding = DingtalkChatbot(webhook, pc_slide=True)  # 方式三：设置消息链接在PC端侧边栏打开（v1.5以上新功能）
    # Text消息@所有人
    # xiaoding.send_text(msg=c, is_at_all=True)

    # fscan扫描漏洞信息
    new_list4 = []
    num4 = len(fs_list)
    print(fs_list)
    if int(num4) > 100:
        n4 = num4 // 50
    else:
        n4 = 1

    for i in range(n4):
        one_list = fs_list[math.floor(i / n4 * num4):math.floor((i + 1) / n4 * num4)]
        new_list4.append(one_list)
    if num4 != 0:
        for i in new_list4:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集fscan扫描漏洞信息 ' + str(
                num4) + ' 个' + '\n' + '-----------------------------------------------'

            for msg in i:
                message = message + str(xuhao) + '.' + str(msg) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg4 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg4)

    #漏洞信息个数
    msg_info = []
    # 漏洞信息
    new_list3 = []
    num3 = len(ld_list)
    print(ld_list)
    if int(num3) > 100:
        n3 = num3 // 50
    else:
        n3 = 1

    for i in range(n3):
        one_list = ld_list[math.floor(i / n3 * num3):math.floor((i + 1) / n3 * num3)]
        new_list3.append(one_list)
    if num3 != 0:
        for i in new_list3:
            xuhao = 1
            message = ''
            title = dingding_tag + ':新收集漏洞信息 ' + str(
                num3) + ' 个,排除一些抽象漏洞后如下' + '\n' + '-----------------------------------------------'

            for msg in i:
                msg_info.append(msg[0])

            for msg in i:
                if 'ssl' in msg[0] or 'tls' in msg[0] or '-certificate' in msg[0] or 'cipher' in msg[0] or 'mismatched' in msg[0] or msg_info.count(msg[0]) > 6:
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2])
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg3 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg3)

    # 敏感数据
    new_list2 = []
    num2 = len(mgml_list)
    print(mgml_list)
    if int(num2) > 100:
        n2 = num2 // 50
    else:
        n2 = 1

    for i in range(n2):
        one_list = mgml_list[math.floor(i / n2 * num2):math.floor((i + 1) / n2 * num2)]
        new_list2.append(one_list)
    if num2 != 0:
        for i in new_list2:
            xuhao = 1
            message = ''
            title = dingding_tag+':新收集敏感信息 ' + str(
                num2) + ' 个,其中返回为200的如下' + '\n' + '-----------------------------------------------'

            for msg in i:
                if str(msg[1]) != '200':
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2])
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg2 = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg2)

    # 单独@人员
    new_list = []
    num = len(message_list)
    print(message_list)
    if int(num) > 200:
        n = num // 50
    else:
        n = 1
    for i in range(n):
        one_list = message_list[math.floor(i / n * num):math.floor((i + 1) / n * num)]
        new_list.append(one_list)
    if num != 0:
        for i in new_list:
            print(i)
            xuhao = 1
            num1 = len(i)
            title = dingding_tag+':新收集暴露面信息 ' + str(
                num1) + ' 个,其中状态码为200的如下' + '\n' + '-----------------------------------------------' + '\n' + '网址           ' + '    状态码    ' + '     标题      '
            message = ''
            for msg in i:
                if str(msg[1]) != '200':
                    continue
                info = str(msg[0]) + '   ' + str(msg[1]) + '   ' + str(msg[2]) + '   '
                message = message + str(xuhao) + '.' + str(info) + '\n'
                xuhao += 1
                message = title.lstrip() + '\n' + message
                title = ''
            message + '\n' + '-----------------------------------------------'
            msg = message.lstrip('\n')
            if message != '':
                xiaoding.send_text(msg=msg)


# 调用搜集的所有IP列表进行fscan扫描
def fscan(filename,ip_list):
    l = open('./result/allip/ip_cache.txt', 'r', encoding='utf-8').read().split('\n')
    url_file = './result/allip/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_allip.txt'
    loud_file = './result/fscan/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_fscan.txt'
    loud_file2 = './result/fscan/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + '_ip_fscan.txt'
    with open(url_file,'w') as f:
        for ip in ip_list:
            if ip not in l:
                f.writelines(ip + '\n')
    os.system('./inifile/lousao/fscan  -uf ' + filename + ' -o ' + str(loud_file))
    #os.system('./inifile/lousao/fscan  -p 22,3389,445,3306,1433,1521,21,27017,11211,5432,23,25,465,110,995,143,993,5900,6379 -np -hf ' + url_file + ' -o ' + str(loud_file2))
    os.system('./inifile/lousao/fscan  -np -hf ' + url_file + ' -o ' + str(loud_file2))
    list1 = []
    with open('./result/allip/ip_cache.txt','a') as f:
        for ip in ip_list:
            f.writelines(ip + '\n')
    try:
        with open(loud_file, 'r') as f:
            print(loud_file)
            test1 = f.readlines()
            for t in test1:
                if '[+]' in t and '扫描结束' not in t:
                    list1.append(t.strip('\n'))
    except:
        print('无漏洞')
    try:
        with open(loud_file2, 'r') as f:
            print(loud_file2)
            test1 = f.readlines()
            for t in test1:
                if '[+]' in t and '扫描结束' not in t:
                    list1.append(t.strip('\n'))
    except:
        print('无漏洞')
    # try:
    #     with open(loud_file, 'r') as f:
    #         print(loud_file)
    #         test1 = f.readlines()
    #         for t in test1:
    #             if '[*]' in t and 'alive ports' not in t:
    #                 list1.append(t.strip('\n'))
    # except:
    #     print('无漏洞!')
    # try:
    #     with open(loud_file2, 'r') as f:
    #         print(loud_file2)
    #         test1 = f.readlines()
    #         for t in test1:
    #             if '[*]' in t and 'alive ports' not in t:
    #                 list1.append(t.strip('\n'))
    # except:
    #     print('无漏洞!')

    print(list1)
    return list1

def nuclei(filename):
    os.system('./inifile/lousao/nuclei -update')
    os.system('./inifile/lousao/nuclei -update-templates ')
    loud_file = './result/loudong/' + time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) + 'ld_scan.txt'
    # os.system('./inifile/lousao/nuclei -un -ut')
    os.system('./inifile/lousao/nuclei -mhe 3 -timeout 1 -rl 300 -c 50  -s low,medium,high,critical -l ' + str(filename) + ' -o ' + str(loud_file))
    list1 = []
    list2 = []
    try:
        with open(loud_file, 'r') as f:
            print(loud_file)
            test1 = f.readlines()
            for t in test1:
                if t[0] != '#' and t[0] != '\n':
                    list1.append(t.strip('\n'))
    except:
        print('该网站无漏洞信息')
        return list2

    for l in list1:
        temp = []
        x = l.split(' ')
        print(x)
        temp.append(x[0])
        temp.append(x[2])
        temp.append(x[3])
        list2.append(temp)
    print(list2)

    return list2


def get_github_info(company_info_list, all_company_name_list):
    name_list = []
    info_list = []
    for com in company_info_list:
        name_list.append(com[0])
    for com in all_company_name_list:
        name_list.append(com)

    for name in name_list:
        page = 1
        while True:
            header = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                      "Cookie": "",
                      "X-CSRFToken": ""}
            title = '(related_company==' + str(name) + '||url==' + str(name) + '||repository.description==' + str(
                name) + '||code_detail==' + str(name) + ')'
            data = 'page=' + str(page) + '&pagesize=10&title=' + str(title) + '&title_type=code'
            print(data)
            proxies = {'http': 'http://localhost:8080', 'https': 'http://localhost:8080'}
            a = requests.post('https://0.zone/api/home/search/', data=data.encode('utf-8'), headers=header,
                              verify=False, proxies=proxies).json()
            if str(a['code']) == '1':
                continue
            if str(a['code']) == '3':
                break
            res = a['data']['data_list']
            for r in res:
                info = []
                tag = r['_source']['tags']
                type = r['_source']['type']
                tags = ''
                for t in tag:
                    tags = tags + t + ' || '
                tags = tags + '||' + type
                source = r['_source']['source']
                url = r['_source']['url']
                code_detail = r['_source']['code_detail']
                info.append(tags)
                info.append(url)
                info.append(code_detail)
                info.append(source)
                info_list.append(info)
            page += 1

    return info_list

def quchong(l1):
    temp_list = []
    l = []
    for i in l1:
        try:
            temp_list.append(','.join(i))
        except:
            print('jon')
            continue
    print(temp_list)
    l2 = list(set(temp_list))
    print(l2)
    for x in l2:
        l.append(x.split(','))

    return l

if __name__ == '__main__':
    # 参数设置
    parser = optparse.OptionParser()
    parser.add_option('-c', '--company', action='store', dest="company_name")
    parser.add_option('-x', '--xdomain', action='store', dest="x_domain")
    parser.add_option('-l', '--list', action='store', dest="company_name_list")
    parser.add_option('-o', '--occ', action='store', default='100', dest="company_occ")
    parser.add_option('-d', '--domain', action='store', default='', dest="company_domains")
    parser.add_option('-z', '--zs', action='store_true', dest="zs_domains")
    parser.add_option('-r', '--recursion', action='store', default='', dest="recursion_level")
    parser.add_option('-m', '--ml', action='store_true', dest="ml_sm")
    parser.add_option('-n', '--nl', action='store_true', dest="ld_sm")
    parser.add_option('-f', '--fs', action='store_true', dest="fs_sm")
    parser.add_option('-a', '--av', action='store_true', dest="av_sm")
    parser.add_option('-N', '--notauto', action='store_true', dest="not_auto")
    options, args = parser.parse_args()

    # 加载配置文件
    cf = ConfigParser()
    cf.read('./config/config.ini')
    global hunter_config_list
    hunter_config_list = []
    global fofa_key
    fofa_key = ''
    global fofa_email
    fofa_email = ''
    global black_domian
    black_domian = []
    global dingding_key
    dingding_key = ''
    global dingding_hook
    dingding_hook = ''
    global x_auth_token
    x_auth_token = ''
    global fofa_size
    fofa_size = ''
    global fofa_keyword
    fofa_keyword = ''
    global yt_size
    yt_size = ''
    global yt_keword
    yt_keword = ''
    global fofa_count
    fofa_count = ''
    global hunter_count
    hunter_count = ''
    global is_fofa
    is_fofa = ''
    global is_hunter
    is_hunter = ''
    global file_filter_name
    file_filter_name = ''
    global httpx_info
    httpx_info = []
    global dingding_tag
    dingding_tag = ''

    c_len = cf.options('hunter')
    for i in c_len:
        hunter_config_list.append(cf.get('hunter', i))
    yt_keword = ''
    fofa_count = cf.get('fofa', 'count')
    hunter_count = cf.get('hunter', 'count')
    fofa_size = cf.get('fofa', 'size')
    is_fofa = cf.get('fofa', 'is_fofa')
    is_hunter = cf.get('hunter', 'is_hunter')
    yt_size = cf.get('hunter', 'size')
    fofa_keyword = cf.get('fofa', 'keyword')
    yt_keword = cf.get('hunter', 'keyword')
    fofa_key = cf.get('fofa', 'fofa_key')
    fofa_email = cf.get('fofa', 'fofa_email')
    dingding_hook = cf.get('dingding', 'access_token')
    dingding_key = cf.get('dingding', 'dsecret')
    x_auth_token = cf.get('tyz', 'x-auth-token')
    black = cf.get('black_domain', 'domain')
    black_domian = black.split(',')
    dingding_tag = cf.get('tag','dingding_tag')
    # 所有搜集到的URL列表
    all_url_list = []

    # 最后整理出的url列表
    all_url_list2 = []

    # 全局收集的信息
    global all_info_list
    all_info_list = []

    # 全局搜寻的域名
    all_domain_list = []

    # 初始化
    quchong_list = []
    mgwj_list = []
    ld_list = []
    fs_list = []
    global ip_list
    ip_list = []

    # 参数获取
    global company_occ
    company_name = options.company_name
    company_domains_file = options.company_domains
    company_list = options.company_name_list
    company_occ = float(options.company_occ)

    global all_company_name_list
    all_company_name_list = []

    global ml
    ml = options.ml_sm

    global ld
    ld = options.ld_sm

    global fs
    fs = options.fs_sm

    global avsm
    avsm = options.av_sm

    global notauto
    notauto = options.not_auto

    global zs_domains
    zs_domain = options.zs_domains

    global x_domain
    x_domain = options.x_domain


    if zs_domain != True:
        company_id_name_list = Get_Company_Id(company_name, company_list)
        print(company_id_name_list)
        company_info_list, company_url_list = Get_ALL_Sub_Cmpany_Domain(company_id_name_list)
    else:
        company_info_list = []

    # 调用fofa,yt获取信息
    get_all_url_fo_yt(company_info_list, company_domains_file)
    quchong_list, mgwj_list, ld_list,fs_list = quchong_info_list(all_info_list)

    # github监控
    print(all_company_name_list)
    # github_list = get_github_info(company_info_list,all_company_name_list)
    github_list = []
    if len(quchong_list) != 0 or len(github_list) != 0 or len(mgwj_list) != 0 or len(ld_list) != 0 or len(fs_list) != 0:
        Write_To_Excel(company_info_list, quchong_list, mgwj_list, ld_list, httpx_info, fs_list)
        # 发送信息
        try:
            set_info = quchong(httpx_info)
            dingtalk(set_info, mgwj_list, ld_list,fs_list)
        except:
            print('发送消息异常')
            traceback.print_exc()
            time.sleep(60)
            try:
                dingtalk(set_info, mgwj_list, ld_list, fs_list)
            except:
                print('网络存在问题,继续执行任务')

    if notauto != True:
        time.sleep(360)
        try:
            os.system('rm -rf laoyue.out')
            os.system('nohup python3 laoyue.py -d "SRC.txt" -z -n -m -f -a > laoyue.out 2>&1 &')
        except:
            print('laoyue.out文件不存在')
        # nohup ./check_nohup_size.sh >check_size.out 2>&1 & (运行该命令之前,请先运行build.sh文件或者手动在shell执行命令:sed -i "s/\r//" check_nohup_size.sh,定时检查nohup.out是否变化防止卡死导致自动化停止)
